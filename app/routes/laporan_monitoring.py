# app/routes/laporan_monitoring.py
from flask import Blueprint, request, jsonify, send_file, current_app
from datetime import datetime, date, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo
import os
import uuid
import re

from flask_jwt_extended import jwt_required, get_jwt
from sqlalchemy import func
from werkzeug.utils import secure_filename
from app.extensions import db

from app.models.monitoring import LaporanMonitoring
from app.models.mengajar import LaporanMengajar
from app.models.kehadiran_guru import KehadiranGuru
from app.models.jadwal import Jadwal
from app.models.jadwal_guru import JadwalGuru
from app.models.kelas import Kelas
from app.models.tingkat import Tingkat
from app.models.mata_pelajaran import MataPelajaran
from app.models.guru import Guru
from app.models.murid import Murid
from app.models.kehadiran_murid import KehadiranMurid
from app.models.periode_akademik import PeriodeAkademik

monitoring_bp = Blueprint("monitoring", __name__)


# =====================================================
# TIMEZONE HELPER
# =====================================================
# Atur zona waktu aplikasi dari .env sesuai lokasi sekolah/instansi.
# Contoh:
# APP_TIMEZONE=Asia/Jakarta   -> WIB
# APP_TIMEZONE=Asia/Makassar  -> WITA
# APP_TIMEZONE=Asia/Jayapura  -> WIT
_DEFAULT_APP_TIMEZONE = "Asia/Jakarta"


def _app_timezone():
    tz_name = os.getenv("APP_TIMEZONE", _DEFAULT_APP_TIMEZONE).strip() or _DEFAULT_APP_TIMEZONE
    try:
        return ZoneInfo(tz_name)
    except Exception:
        return ZoneInfo(_DEFAULT_APP_TIMEZONE)


def _now_app():
    return datetime.now(_app_timezone())


def _now_app_naive():
    # Database MySQL umumnya menyimpan DATETIME tanpa timezone.
    # Karena itu timezone lokal aplikasi dibuang sebelum disimpan.
    return _now_app().replace(tzinfo=None)


def _today_app():
    return _now_app().date()


# =====================================================
# HELPER
# =====================================================
def jadwal_milik_guru(id_jadwal: int, id_guru: int) -> bool:
    return db.session.query(JadwalGuru).filter_by(
        id_jadwal=id_jadwal,
        id_guru=id_guru
    ).first() is not None


def hari_indonesia_lower():
    map_hari = {
        "Monday": "senin",
        "Tuesday": "selasa",
        "Wednesday": "rabu",
        "Thursday": "kamis",
        "Friday": "jumat",
        "Saturday": "sabtu",
        "Sunday": "minggu",
    }
    return map_hari[_now_app().strftime("%A")]


def _fmt_time(value):
    return value.strftime("%H:%M:%S") if value else None


def _fmt_jadwal_time(value):
    return value.strftime("%H:%M") if value else None


def _safe_sheet_value(value):
    if value is None:
        return "-"
    text = str(value)
    if text.strip() == "" or text.strip().lower() == "none":
        return "-"
    return text


def get_tingkat_info(kelas):
    if not kelas:
        return None, "-"

    id_tingkat = getattr(kelas, "id_tingkat", None)
    tingkat = getattr(kelas, "tingkat", None)

    if tingkat is None and id_tingkat:
        tingkat = Tingkat.query.get(id_tingkat)

    pangkat = None
    if tingkat is not None:
        pangkat = (
            getattr(tingkat, "pangkat", None)
            or getattr(tingkat, "nama_tingkat", None)
            or getattr(tingkat, "tingkat", None)
        )

    tingkat_text = str(pangkat) if pangkat is not None else (str(id_tingkat) if id_tingkat is not None else "-")
    return id_tingkat, tingkat_text




def _status_text(value, default="aktif"):
    text = str(value if value is not None else default).strip().lower()
    return text or default


_STATUS_TIDAK_AKTIF = [
    "selesai", "arsip", "diarsipkan", "archive", "archived",
    "nonaktif", "non aktif", "non-aktif", "inactive",
]


def _status_belum_selesai_expr(column):
    """Filter SQL: status NULL/kosong dianggap aktif; status selesai/arsip/nonaktif tidak ditampilkan."""
    return ~func.lower(func.trim(func.coalesce(column, "aktif"))).in_(_STATUS_TIDAK_AKTIF)


def _jadwal_kelas_belum_selesai_expr():
    return db.and_(
        _status_belum_selesai_expr(Jadwal.status),
        _status_belum_selesai_expr(Kelas.status),
    )


def _jadwal_kelas_selesai(jadwal):
    if not jadwal:
        return True
    if _status_text(getattr(jadwal, "status", None)) in _STATUS_TIDAK_AKTIF:
        return True

    kelas = getattr(jadwal, "kelas", None)
    if kelas is None and getattr(jadwal, "id_kelas", None):
        kelas = Kelas.query.get(jadwal.id_kelas)

    return _status_text(getattr(kelas, "status", None)) in _STATUS_TIDAK_AKTIF

def _hapus_monitoring_lebih_14_hari():
    """
    Riwayat monitoring hanya disimpan selama 14 hari.
    Data laporan_monitoring yang lebih lama akan dihapus bersama laporan_mengajar terkait.
    """
    cutoff = _today_app() - timedelta(days=14)

    old_ids = [
        row[0]
        for row in db.session.query(LaporanMonitoring.id_monitor)
        .filter(LaporanMonitoring.tanggal < cutoff)
        .all()
    ]

    if not old_ids:
        return 0

    LaporanMengajar.query.filter(
        LaporanMengajar.id_monitor.in_(old_ids)
    ).delete(synchronize_session=False)

    LaporanMonitoring.query.filter(
        LaporanMonitoring.id_monitor.in_(old_ids)
    ).delete(synchronize_session=False)

    db.session.commit()
    return len(old_ids)


def _laporan_payload(laporan):
    if not laporan:
        return None

    return {
        "id_laporan": laporan.id_laporan,
        "id_monitor": laporan.id_monitor,
        "materi": laporan.materi,
        "catatan": laporan.catatan,
        "jumlah_hadir": laporan.jumlah_hadir,
        "jumlah_tidak_hadir": getattr(laporan, "jumlah_tidak_hadir", 0),
        "bawa_data_kehadiran": bool(getattr(laporan, "bawa_data_kehadiran", False)),
        "daftar_hadir": getattr(laporan, "daftar_hadir", None),
        "daftar_tidak_hadir": getattr(laporan, "daftar_tidak_hadir", None),
        "waktu_input": laporan.waktu_input.strftime("%Y-%m-%d %H:%M:%S")
        if laporan.waktu_input else None,
    }


def _murid_label(murid):
    nis = getattr(murid, "nis", None) or getattr(murid, "nisn", None) or "-"
    nama = getattr(murid, "nama_murid", None) or getattr(murid, "nama", None) or "-"
    return f"{nis} - {nama}"


def _jadwal_group_ids_absensi(id_jadwal):
    """
    Ambil semua id_jadwal yang masih satu kelas + mapel.
    Dipakai khusus laporan mengajar agar tombol "membawa data kehadiran"
    tetap bisa membaca absensi dari Input Kehadiran walaupun mapel punya J1/J2/J3.
    """
    jadwal = Jadwal.query.get(id_jadwal)
    if not jadwal:
        return None, []

    id_kelas = getattr(jadwal, "id_kelas", None)
    id_mapel = getattr(jadwal, "id_mapel", None)

    if id_kelas is None or id_mapel is None:
        return jadwal, [id_jadwal]

    rows = (
        Jadwal.query
        .filter(
            Jadwal.id_kelas == id_kelas,
            Jadwal.id_mapel == id_mapel,
        )
        .order_by(Jadwal.jam_mulai.asc(), Jadwal.id_jadwal.asc())
        .all()
    )

    ids = [j.id_jadwal for j in rows if getattr(j, "id_jadwal", None) is not None]
    return jadwal, ids or [id_jadwal]


def _periode_absensi_aktif_values():
    periode = PeriodeAkademik.aktif()
    if not periode:
        return None, None
    return periode.semester, periode.tahun_ajaran


def _absensi_murid_query(id_jadwal_group, pertemuan=None):
    """
    Query data kehadiran pada grup jadwal mapel yang sama.

    Data difilter memakai periode akademik aktif agar pertemuan lama dari
    tahun ajaran/semester sebelumnya tidak ikut ditampilkan pada laporan.
    """
    query = (
        db.session.query(KehadiranMurid, Murid)
        .join(Murid, Murid.id_murid == KehadiranMurid.id_murid)
        .filter(KehadiranMurid.id_jadwal.in_(id_jadwal_group))
    )

    semester, tahun_ajaran = _periode_absensi_aktif_values()
    if semester:
        query = query.filter(KehadiranMurid.semester == semester)
    if tahun_ajaran:
        query = query.filter(KehadiranMurid.tahun_ajaran == tahun_ajaran)
    if pertemuan is not None:
        query = query.filter(KehadiranMurid.pertemuan == pertemuan)

    return query


def _pertemuan_terisi_absensi(id_jadwal_group):
    if not id_jadwal_group:
        return []

    query = db.session.query(KehadiranMurid.pertemuan).filter(
        KehadiranMurid.id_jadwal.in_(id_jadwal_group)
    )
    semester, tahun_ajaran = _periode_absensi_aktif_values()
    if semester:
        query = query.filter(KehadiranMurid.semester == semester)
    if tahun_ajaran:
        query = query.filter(KehadiranMurid.tahun_ajaran == tahun_ajaran)

    rows = (
        query
        .filter(KehadiranMurid.pertemuan.isnot(None))
        .distinct()
        .order_by(KehadiranMurid.pertemuan.asc())
        .all()
    )
    return [int(row[0]) for row in rows if row[0] is not None]


def _absensi_murid_payload(id_jadwal, pertemuan=None):
    """
    Ambil data kehadiran berdasarkan pertemuan yang sudah diinput.

    Sebelumnya data dicari berdasarkan tanggal laporan/hari ini. Cara itu dapat
    menghasilkan data kosong ketika tanggal input absensi berbeda, timezone
    server berbeda, atau guru membuka laporan setelah pergantian tanggal.
    Sekarang sumber data dipilih berdasarkan nomor pertemuan pada periode aktif.
    """
    jadwal, id_jadwal_group = _jadwal_group_ids_absensi(id_jadwal)

    total_murid = 0
    if jadwal:
        total_murid = Murid.query.filter_by(id_kelas=jadwal.id_kelas).count()

    pertemuan_terisi = _pertemuan_terisi_absensi(id_jadwal_group)

    if pertemuan is not None:
        try:
            pertemuan = int(pertemuan)
        except Exception:
            pertemuan = None

    # Kompatibilitas dengan frontend lama: bila belum mengirim pertemuan,
    # gunakan pertemuan terisi paling akhir. Frontend baru tetap meminta guru
    # memilih pertemuan secara eksplisit.
    if pertemuan is None and pertemuan_terisi:
        pertemuan = pertemuan_terisi[-1]

    base_payload = {
        "id_jadwal": id_jadwal,
        "id_jadwal_group": id_jadwal_group,
        "pertemuan": pertemuan,
        "pertemuan_terisi": pertemuan_terisi,
        "total_murid": total_murid,
        "jumlah_hadir": 0,
        "jumlah_tidak_hadir": 0,
        "daftar_hadir": "",
        "daftar_tidak_hadir": "",
        "data_tersedia": False,
    }

    if not id_jadwal_group or pertemuan is None or pertemuan not in pertemuan_terisi:
        return base_payload

    rows = (
        _absensi_murid_query(id_jadwal_group, pertemuan=pertemuan)
        .order_by(KehadiranMurid.id_kehadiran.asc())
        .all()
    )

    # Antisipasi data lama yang mungkin tersimpan lebih dari sekali pada grup
    # J1/J2/J3: ambil data terakhir per murid.
    latest = {}
    for kehadiran, murid in rows:
        old = latest.get(kehadiran.id_murid)
        if old is None or kehadiran.id_kehadiran > old[0].id_kehadiran:
            latest[kehadiran.id_murid] = (kehadiran, murid)

    hadir = []
    tidak_hadir = []

    for kehadiran, murid in latest.values():
        status = str(kehadiran.status or "").strip().lower()
        label = _murid_label(murid)
        if status == "hadir":
            hadir.append(label)
        else:
            status_label = kehadiran.status or "Alpa"
            tidak_hadir.append(f"{label} ({status_label})")

    base_payload.update({
        "jumlah_hadir": len(hadir),
        "jumlah_tidak_hadir": len(tidak_hadir),
        "daftar_hadir": "; ".join(hadir),
        "daftar_tidak_hadir": "; ".join(tidak_hadir),
        "data_tersedia": len(latest) > 0,
    })
    return base_payload

def _get_kehadiran_guru(id_guru, tanggal, id_jadwal=None):
    if not id_guru or not tanggal:
        return None

    query = KehadiranGuru.query.filter_by(
        id_guru=id_guru,
        tanggal=tanggal
    )

    # Jika id_jadwal dikirim, status kehadiran harus spesifik per jadwal/mapel.
    # Ini mencegah 1 status hadir guru dipakai untuk semua mapel di tanggal yang sama.
    if id_jadwal is not None:
        query = query.filter_by(id_jadwal=id_jadwal)

    return query.first()


def _upsert_kehadiran_guru(
    id_guru,
    tanggal,
    id_jadwal=None,
    status="Hadir",
    keterangan=None,
    alasan=None,
    instruksi=None,
    bukti=None,
    status_pengajuan=None,
):
    """
    Rekap kehadiran guru disimpan per jadwal.
    Status izin/sakit juga disimpan di tabel yang sama agar admin bisa melihat
    pengajuan tanpa harus menunggu laporan_monitoring dibuat.
    """
    if not id_guru or not tanggal:
        return None

    item = _get_kehadiran_guru(id_guru, tanggal, id_jadwal)
    next_status = status or (item.status if item else None) or "Hadir"

    if item:
        item.status = next_status

        if keterangan:
            old = (item.keterangan or "").strip()
            if keterangan not in old:
                item.keterangan = f"{old}; {keterangan}" if old else keterangan

        if alasan is not None:
            item.alasan = alasan
        if instruksi is not None:
            item.instruksi = instruksi
        if bukti is not None:
            item.bukti = bukti
        if status_pengajuan is not None:
            item.status_pengajuan = status_pengajuan

        if str(next_status).strip().lower() == "hadir":
            item.alasan = None
            item.instruksi = None
            item.bukti = None
            item.status_pengajuan = None

        return item

    item = KehadiranGuru(
        id_guru=id_guru,
        id_jadwal=id_jadwal,
        tanggal=tanggal,
        status=next_status,
        keterangan=keterangan,
        alasan=alasan,
        instruksi=instruksi,
        bukti=bukti,
        status_pengajuan=status_pengajuan,
    )
    db.session.add(item)
    return item


def _jadwal_label(id_jadwal):
    row = (
        db.session.query(Jadwal, Kelas, MataPelajaran)
        .join(Kelas, Kelas.id_kelas == Jadwal.id_kelas)
        .join(MataPelajaran, MataPelajaran.id_mapel == Jadwal.id_mapel)
        .filter(Jadwal.id_jadwal == id_jadwal)
        .first()
    )

    if not row:
        return f"Jadwal {id_jadwal}"

    jadwal, kelas, mapel = row
    return f"{kelas.nama_kelas} - {mapel.nama_mapel}"

def _allowed_bukti_file(filename):
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext in {"jpg", "jpeg", "png", "webp", "pdf"}


def _save_bukti_file(file_storage):
    if not file_storage or not file_storage.filename:
        return None

    if not _allowed_bukti_file(file_storage.filename):
        raise ValueError("Format bukti harus jpg, jpeg, png, webp, atau pdf")

    filename = secure_filename(file_storage.filename)
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "jpg"
    unique_name = f"{uuid.uuid4().hex}.{ext}"

    upload_dir = os.path.join(current_app.root_path, "static", "uploads", "izin_sakit")
    os.makedirs(upload_dir, exist_ok=True)

    file_storage.save(os.path.join(upload_dir, unique_name))
    return f"static/uploads/izin_sakit/{unique_name}"


def _bukti_url(path):
    if not path:
        return None

    text = str(path).strip()
    if not text or text.lower() == "none":
        return None

    if text.startswith("http://") or text.startswith("https://"):
        return text

    base = request.host_url.rstrip("/") if request else ""
    return f"{base}/{text.lstrip('/')}"


def _add_keterangan(item, text):
    if not item or not text:
        return
    old = (item.keterangan or "").strip()
    if text not in old:
        item.keterangan = f"{old}; {text}" if old else text


def _waktu_dari_keterangan(keterangan, jenis):
    text = str(keterangan or "")
    if jenis == "pengajuan":
        pattern = r"Pengajuan (?:Izin|Sakit) dikirim (\d{2}:\d{2}:\d{2})"
    else:
        pattern = r"(?:disetujui|ditolak) admin (\d{2}:\d{2}:\d{2})"

    matches = re.findall(pattern, text, flags=re.IGNORECASE)
    return matches[-1] if matches else None


def _pengajuan_value(item):
    return str(getattr(item, "status_pengajuan", "") or "").strip().lower()


def _status_kehadiran_manual(kehadiran_guru):
    if not kehadiran_guru or not kehadiran_guru.status:
        return None

    text = str(kehadiran_guru.status).strip().lower()
    if text in ["izin", "ijin"]:
        return "Izin"
    if text == "sakit":
        return "Sakit"
    if text in ["alpa", "alpha", "tidak hadir", "tidak_hadir"]:
        return "Alpa"
    if text in ["hadir", "masuk", "selesai"]:
        return "Hadir"
    return kehadiran_guru.status


def _jadwal_sudah_selesai(jadwal, tanggal_value):
    if not jadwal or not jadwal.jam_selesai or not tanggal_value:
        return False

    today = _today_app()
    if tanggal_value < today:
        return True
    if tanggal_value > today:
        return False

    return _now_app().time() > jadwal.jam_selesai


def _status_monitoring(jadwal, monitor=None, laporan=None, kehadiran_guru=None, tanggal_override=None):
    """
    Status final yang dipakai admin monitoring.
    Urutan status yang dipakai hanya: Selesai, Hadir, Alpa, Izin, Belum Absen.
    Patokan alpa bukan lagi hari saja, tapi jam_selesai jadwal.
    """
    tanggal_value = tanggal_override or (monitor.tanggal if monitor and monitor.tanggal else _today_app())

    if monitor and monitor.jam_keluar:
        return "Selesai"

    if monitor and monitor.jam_masuk:
        return "Hadir"

    pengajuan = _pengajuan_value(kehadiran_guru)
    manual = _status_kehadiran_manual(kehadiran_guru)

    if manual in ["Izin", "Sakit"]:
        if pengajuan == "ditolak":
            return "Alpa" if _jadwal_sudah_selesai(jadwal, tanggal_value) else "Belum Absen"
        return manual

    if manual == "Alpa":
        if pengajuan == "ditolak" and not _jadwal_sudah_selesai(jadwal, tanggal_value):
            return "Belum Absen"
        return "Alpa"

    if manual == "Hadir":
        return "Hadir"

    if _jadwal_sudah_selesai(jadwal, tanggal_value):
        return "Alpa"

    return "Belum Absen"


def _monitoring_payload(jadwal, kelas, mapel, guru, monitor=None, laporan=None, kehadiran_guru=None, tanggal_override=None):
    masuk = None
    keluar = None
    id_monitor = None
    tanggal_value = _today_app()
    id_tingkat, tingkat_text = get_tingkat_info(kelas)

    if tanggal_override is not None:
        tanggal_value = tanggal_override

    if monitor:
        id_monitor = monitor.id_monitor
        tanggal_value = monitor.tanggal
        masuk = _fmt_time(monitor.jam_masuk)
        keluar = _fmt_time(monitor.jam_keluar)

    if kehadiran_guru is None:
        kehadiran_guru = _get_kehadiran_guru(
            guru.id_guru if guru else None,
            tanggal_value,
            jadwal.id_jadwal if jadwal else None
        )

    status = _status_monitoring(jadwal, monitor, laporan, kehadiran_guru, tanggal_override=tanggal_value)

    return {
        "id": id_monitor,
        "id_monitor": id_monitor,
        "tanggal": str(tanggal_value) if tanggal_value else str(_today_app()),
        "id_jadwal": jadwal.id_jadwal,
        "id_tingkat": id_tingkat,
        "tingkat": tingkat_text,
        "pangkat": tingkat_text,

        "id_guru": guru.id_guru if guru else None,
        "nip": getattr(guru, "nip", None) if guru else None,
        "nama_guru": guru.nama_guru if guru else "-",
        "guru": guru.nama_guru if guru else "-",

        "kelas": kelas.nama_kelas if kelas else "-",
        "mapel": mapel.nama_mapel if mapel else "-",
        "status_jadwal": getattr(jadwal, "status", "aktif") if jadwal else "aktif",
        "status_kelas": getattr(kelas, "status", "aktif") if kelas else "aktif",
        "hari_jadwal": getattr(jadwal, "hari", None) if jadwal else None,
        "jam_jadwal_mulai": _fmt_jadwal_time(jadwal.jam_mulai),
        "jam_jadwal_selesai": _fmt_jadwal_time(jadwal.jam_selesai),
        "masuk": masuk,
        "keluar": keluar,
        "status": status,

        "id_kehadiran_guru": kehadiran_guru.id_kehadiran if kehadiran_guru else None,
        "kehadiran_guru": kehadiran_guru.status if kehadiran_guru else None,
        "keterangan_kehadiran": kehadiran_guru.keterangan if kehadiran_guru else None,
        "alasan": getattr(kehadiran_guru, "alasan", None) if kehadiran_guru else None,
        "instruksi": getattr(kehadiran_guru, "instruksi", None) if kehadiran_guru else None,
        "bukti": getattr(kehadiran_guru, "bukti", None) if kehadiran_guru else None,
        "bukti_url": _bukti_url(getattr(kehadiran_guru, "bukti", None)) if kehadiran_guru else None,
        "status_pengajuan": getattr(kehadiran_guru, "status_pengajuan", None) if kehadiran_guru else None,

        "sudah_laporan": laporan is not None,
        "laporan_mengajar": _laporan_payload(laporan),
        "materi": laporan.materi if laporan else None,
        "catatan": laporan.catatan if laporan else None,
        "jumlah_hadir": laporan.jumlah_hadir if laporan else None,
        "jumlah_tidak_hadir": getattr(laporan, "jumlah_tidak_hadir", None) if laporan else None,
        "bawa_data_kehadiran": bool(getattr(laporan, "bawa_data_kehadiran", False)) if laporan else False,
        "daftar_hadir": getattr(laporan, "daftar_hadir", None) if laporan else None,
        "daftar_tidak_hadir": getattr(laporan, "daftar_tidak_hadir", None) if laporan else None,
        "waktu_input_laporan": laporan.waktu_input.strftime("%Y-%m-%d %H:%M:%S")
        if laporan and laporan.waktu_input else None,
    }


def _unpack_monitoring_row(row):
    if len(row) >= 7:
        jadwal, kelas, mapel, guru, monitor, laporan, kehadiran = row[:7]
        return jadwal, kelas, mapel, guru, monitor, laporan, kehadiran

    jadwal, kelas, mapel, guru, monitor, laporan = row
    return jadwal, kelas, mapel, guru, monitor, laporan, None


def _monitoring_payload_from_row(row):
    jadwal, kelas, mapel, guru, monitor, laporan, kehadiran = _unpack_monitoring_row(row)
    tanggal_override = getattr(kehadiran, "tanggal", None) if kehadiran and not monitor else None
    return _monitoring_payload(
        jadwal,
        kelas,
        mapel,
        guru,
        monitor,
        laporan,
        kehadiran_guru=kehadiran,
        tanggal_override=tanggal_override,
    )


def _sinkron_kehadiran_guru_terjadwal(rows):
    """
    Pastikan data kehadiran_guru tersimpan per jadwal/mapel.
    Jika guru hadir pada mapel A tetapi tidak hadir pada mapel B di hari yang sama,
    database akan punya dua baris: A=Hadir dan B=Alpa setelah waktu jadwal B lewat.
    """
    changed = False

    for row in rows:
        jadwal, kelas, mapel, guru, monitor, laporan, kehadiran_row = _unpack_monitoring_row(row)
        if not jadwal or not guru:
            continue

        tanggal_value = monitor.tanggal if monitor and monitor.tanggal else _today_app()
        item = kehadiran_row or _get_kehadiran_guru(guru.id_guru, tanggal_value, jadwal.id_jadwal)

        if monitor and monitor.jam_masuk:
            desired_status = "Hadir"
            ket = f"Sinkron monitoring hadir - {_jadwal_label(jadwal.id_jadwal)}"
        elif _jadwal_sudah_selesai(jadwal, tanggal_value):
            desired_status = "Alpa"
            ket = f"Auto alpa setelah jadwal selesai - {_jadwal_label(jadwal.id_jadwal)}"
        else:
            continue

        if item is None:
            _upsert_kehadiran_guru(
                id_guru=guru.id_guru,
                tanggal=tanggal_value,
                id_jadwal=jadwal.id_jadwal,
                status=desired_status,
                keterangan=ket
            )
            changed = True
            continue

        current = str(item.status or "").strip().lower()

        # Hadir dari monitoring selalu menguatkan status hadir.
        if desired_status == "Hadir" and current not in ["hadir", "masuk", "selesai"]:
            item.status = "Hadir"
            old = (item.keterangan or "").strip()
            item.keterangan = f"{old}; {ket}" if old and ket not in old else (old or ket)
            changed = True

        # Auto alpa tidak menimpa izin/sakit yang sudah disetujui.
        # Jika pengajuan ditolak dan jadwal sudah lewat, baru berubah menjadi Alpa.
        if desired_status == "Alpa" and (
            current in ["", "belum absen", "belum_absen"]
            or _pengajuan_value(item) == "ditolak"
        ):
            item.status = "Alpa"
            old = (item.keterangan or "").strip()
            item.keterangan = f"{old}; {ket}" if old and ket not in old else (old or ket)
            changed = True

    if changed:
        db.session.commit()


def _query_monitoring_rows(mode="today", tanggal_from=None, tanggal_to=None):
    if mode == "history":
        q = (
            db.session.query(
                Jadwal,
                Kelas,
                MataPelajaran,
                Guru,
                LaporanMonitoring,
                LaporanMengajar,
                KehadiranGuru,
            )
            .join(Jadwal, Jadwal.id_jadwal == KehadiranGuru.id_jadwal)
            .join(Guru, Guru.id_guru == KehadiranGuru.id_guru)
            .join(Kelas, Kelas.id_kelas == Jadwal.id_kelas)
            .join(MataPelajaran, MataPelajaran.id_mapel == Jadwal.id_mapel)
            .outerjoin(
                LaporanMonitoring,
                db.and_(
                    LaporanMonitoring.id_jadwal == KehadiranGuru.id_jadwal,
                    LaporanMonitoring.tanggal == KehadiranGuru.tanggal,
                ),
            )
            .outerjoin(LaporanMengajar, LaporanMengajar.id_monitor == LaporanMonitoring.id_monitor)
        )

        cutoff_riwayat = _today_app() - timedelta(days=14)
        q = q.filter(KehadiranGuru.tanggal >= cutoff_riwayat)
        q = q.filter(_jadwal_kelas_belum_selesai_expr())
        q = q.filter(KehadiranGuru.status.in_(["Hadir", "Izin", "Sakit", "Alpa"]))

        if tanggal_from:
            try:
                q = q.filter(
                    KehadiranGuru.tanggal >= datetime.strptime(tanggal_from, "%Y-%m-%d").date()
                )
            except Exception:
                pass

        if tanggal_to:
            try:
                q = q.filter(
                    KehadiranGuru.tanggal <= datetime.strptime(tanggal_to, "%Y-%m-%d").date()
                )
            except Exception:
                pass

        return q.order_by(
            KehadiranGuru.tanggal.desc(),
            Jadwal.jam_mulai.desc(),
            Guru.nama_guru.asc(),
        ).all()

    today = _today_app()
    hari = hari_indonesia_lower()

    return (
        db.session.query(
            Jadwal,
            Kelas,
            MataPelajaran,
            Guru,
            LaporanMonitoring,
            LaporanMengajar
        )
        .join(JadwalGuru, JadwalGuru.id_jadwal == Jadwal.id_jadwal)
        .join(Guru, Guru.id_guru == JadwalGuru.id_guru)
        .join(Kelas, Kelas.id_kelas == Jadwal.id_kelas)
        .join(MataPelajaran, MataPelajaran.id_mapel == Jadwal.id_mapel)
        .outerjoin(
            LaporanMonitoring,
            db.and_(
                LaporanMonitoring.id_jadwal == Jadwal.id_jadwal,
                LaporanMonitoring.tanggal == today
            )
        )
        .outerjoin(LaporanMengajar, LaporanMengajar.id_monitor == LaporanMonitoring.id_monitor)
        .filter(
            func.lower(func.trim(Jadwal.hari)) == hari.lower(),
            _jadwal_kelas_belum_selesai_expr(),
        )
        .order_by(Jadwal.jam_mulai.asc(), Guru.nama_guru.asc())
        .all()
    )


# =====================================================
# JADWAL HARI INI (GURU LOGIN)
# =====================================================
# =====================================================
# ABSEN MASUK
# - Simpan detail per jadwal ke laporan_monitoring
# - Simpan rekap harian guru ke kehadiran_guru
# =====================================================
@monitoring_bp.route("/guru/absen-masuk", methods=["POST"])
@jwt_required()
def absen_masuk():
    claims = get_jwt()
    if claims.get("role") != "guru":
        return jsonify({"message": "Akses ditolak"}), 403

    data = request.json or {}
    id_jadwal = data.get("id_jadwal")
    if not id_jadwal:
        return jsonify({"message": "id_jadwal wajib"}), 400

    id_guru = claims.get("id_guru")
    if not id_guru:
        return jsonify({"message": "id_guru tidak ada di token"}), 400

    if not jadwal_milik_guru(int(id_jadwal), int(id_guru)):
        return jsonify({"message": "Jadwal tidak valid"}), 403

    jadwal = Jadwal.query.get_or_404(int(id_jadwal))
    if _jadwal_kelas_selesai(jadwal):
        return jsonify({"message": "Jadwal sudah selesai dan tidak aktif lagi"}), 400

    today = _today_app()

    cek = LaporanMonitoring.query.filter_by(
        id_jadwal=id_jadwal,
        tanggal=today
    ).first()

    if cek:
        return jsonify({
            "message": "Sudah absen",
            "id_monitor": cek.id_monitor
        }), 409

    now_time = _now_app().time()

    monitor = LaporanMonitoring(
        id_jadwal=id_jadwal,
        tanggal=today,
        jam_masuk=now_time,
        status="Hadir"
    )

    db.session.add(monitor)

    _upsert_kehadiran_guru(
        id_guru=id_guru,
        tanggal=today,
        id_jadwal=id_jadwal,
        status="Hadir",
        keterangan=f"Masuk {now_time.strftime('%H:%M:%S')} - {_jadwal_label(id_jadwal)}"
    )

    db.session.commit()

    return jsonify({
        "message": "Absen masuk berhasil",
        "id_monitor": monitor.id_monitor
    }), 201


# =====================================================
# LAPORAN MENGAJAR WAJIB SEBELUM ABSEN KELUAR
# =====================================================
@monitoring_bp.route("/guru/pengajuan-kehadiran", methods=["POST"])
@jwt_required()
def pengajuan_kehadiran_guru():
    claims = get_jwt()
    if claims.get("role") != "guru":
        return jsonify({"message": "Akses ditolak"}), 403

    id_guru = claims.get("id_guru")
    payload_json = request.get_json(silent=True) or {}
    id_jadwal = request.form.get("id_jadwal") or payload_json.get("id_jadwal")
    status = request.form.get("status") or payload_json.get("status")
    alasan = request.form.get("alasan") or payload_json.get("alasan")
    instruksi = request.form.get("instruksi") or payload_json.get("instruksi")

    try:
        id_jadwal = int(id_jadwal)
    except Exception:
        return jsonify({"message": "id_jadwal wajib dan harus valid"}), 400

    status_text = str(status or "").strip().capitalize()
    if status_text not in ["Izin", "Sakit"]:
        return jsonify({"message": "Status pengajuan harus Izin atau Sakit"}), 400

    alasan_text = str(alasan or "").strip()
    if len(alasan_text) < 3:
        return jsonify({"message": "Alasan izin/sakit wajib diisi"}), 400

    instruksi_text = str(instruksi or "").strip()
    if len(instruksi_text) < 3:
        return jsonify({
            "message": "Instruksi yang telah diberikan kepada murid/guru pengganti wajib diisi"
        }), 400

    if not jadwal_milik_guru(id_jadwal, int(id_guru)):
        return jsonify({"message": "Jadwal tidak valid"}), 403

    jadwal = Jadwal.query.get_or_404(id_jadwal)
    if _jadwal_kelas_selesai(jadwal):
        return jsonify({"message": "Jadwal sudah selesai dan tidak aktif lagi"}), 400

    today = _today_app()
    monitor = LaporanMonitoring.query.filter_by(id_jadwal=id_jadwal, tanggal=today).first()
    if monitor and monitor.jam_masuk:
        return jsonify({"message": "Anda sudah absen masuk pada jadwal ini"}), 409

    bukti_path = None
    try:
        bukti_path = _save_bukti_file(request.files.get("bukti"))
    except ValueError as exc:
        return jsonify({"message": str(exc)}), 400

    waktu_pengajuan = _now_app().time().strftime("%H:%M:%S")
    item = _upsert_kehadiran_guru(
        id_guru=id_guru,
        tanggal=today,
        id_jadwal=id_jadwal,
        status=status_text,
        keterangan=(
            f"Pengajuan {status_text} dikirim {waktu_pengajuan} - "
            f"{_jadwal_label(id_jadwal)}"
        ),
        alasan=alasan_text,
        instruksi=instruksi_text,
        bukti=bukti_path,
        status_pengajuan="Menunggu",
    )

    db.session.commit()

    return jsonify({
        "message": f"Pengajuan {status_text.lower()} berhasil dikirim",
        "id_kehadiran_guru": item.id_kehadiran,
        "status": item.status,
        "alasan": item.alasan,
        "instruksi": item.instruksi,
        "bukti": item.bukti,
        "bukti_url": _bukti_url(item.bukti),
        "status_pengajuan": item.status_pengajuan,
        "waktu_pengajuan": waktu_pengajuan,
        "keterangan": item.keterangan,
    }), 201


@monitoring_bp.route("/guru/laporan-mengajar", methods=["POST"])
@jwt_required()
def laporan_mengajar():
    claims = get_jwt()
    if claims.get("role") != "guru":
        return jsonify({"message": "Akses ditolak"}), 403

    data = request.json or {}

    id_monitor = data.get("id_monitor")
    materi = str(data.get("materi") or "").strip()
    catatan = str(data.get("catatan") or "").strip()
    jumlah_hadir = data.get("jumlah_hadir")
    jumlah_tidak_hadir = data.get("jumlah_tidak_hadir")
    bawa_data_kehadiran = str(data.get("bawa_data_kehadiran") or "").strip().lower() in ["true", "1", "ya", "yes"]
    daftar_hadir = str(data.get("daftar_hadir") or "").strip()
    daftar_tidak_hadir = str(data.get("daftar_tidak_hadir") or "").strip()
    pertemuan_kehadiran = data.get("pertemuan_kehadiran") or data.get("pertemuan")

    if not id_monitor:
        return jsonify({"message": "id_monitor wajib"}), 400

    if not materi:
        return jsonify({"message": "Materi wajib diisi"}), 400

    if jumlah_hadir is None or str(jumlah_hadir).strip() == "":
        return jsonify({"message": "Jumlah hadir wajib diisi"}), 400

    try:
        jumlah_hadir = int(jumlah_hadir)
    except Exception:
        return jsonify({"message": "Jumlah hadir harus angka"}), 400

    if jumlah_hadir < 0:
        return jsonify({"message": "Jumlah hadir tidak valid"}), 400

    if jumlah_tidak_hadir is None or str(jumlah_tidak_hadir).strip() == "":
        return jsonify({"message": "Jumlah murid tidak hadir wajib diisi"}), 400

    try:
        jumlah_tidak_hadir = int(jumlah_tidak_hadir)
    except Exception:
        return jsonify({"message": "Jumlah murid tidak hadir harus angka"}), 400

    if jumlah_tidak_hadir < 0:
        return jsonify({"message": "Jumlah murid tidak hadir tidak valid"}), 400

    id_guru = claims.get("id_guru")
    monitor = LaporanMonitoring.query.get_or_404(id_monitor)

    if not jadwal_milik_guru(monitor.id_jadwal, id_guru):
        return jsonify({"message": "Bukan laporan anda"}), 403

    jadwal = Jadwal.query.get(monitor.id_jadwal)
    if _jadwal_kelas_selesai(jadwal):
        return jsonify({"message": "Jadwal sudah selesai dan tidak aktif lagi"}), 400

    if bawa_data_kehadiran:
        if pertemuan_kehadiran is not None and str(pertemuan_kehadiran).strip() != "":
            try:
                pertemuan_kehadiran = int(pertemuan_kehadiran)
            except Exception:
                return jsonify({"message": "Pertemuan kehadiran tidak valid"}), 400

        prefill = _absensi_murid_payload(
            monitor.id_jadwal,
            pertemuan=pertemuan_kehadiran,
        )
        if not prefill["data_tersedia"]:
            return jsonify({
                "message": "Data kehadiran pada pertemuan yang dipilih tidak tersedia",
                "pertemuan_terisi": prefill["pertemuan_terisi"],
            }), 400

        # Selalu hitung ulang dari database agar angka/daftar yang dikirim
        # frontend tidak dapat tertinggal atau berbeda dari Input Kehadiran.
        jumlah_hadir = prefill["jumlah_hadir"]
        jumlah_tidak_hadir = prefill["jumlah_tidak_hadir"]
        daftar_hadir = prefill["daftar_hadir"]
        daftar_tidak_hadir = prefill["daftar_tidak_hadir"]

    if jadwal:
        total_murid = Murid.query.filter_by(id_kelas=jadwal.id_kelas).count()
        if total_murid > 0 and (jumlah_hadir + jumlah_tidak_hadir) != total_murid:
            return jsonify({
                "message": f"Jumlah hadir + tidak hadir harus sama dengan total murid ({total_murid})"
            }), 400

    if monitor.jam_keluar:
        return jsonify({"message": "Laporan tidak bisa diubah setelah absen keluar"}), 400

    laporan = LaporanMengajar.query.filter_by(id_monitor=monitor.id_monitor).first()

    if laporan:
        laporan.materi = materi
        laporan.catatan = catatan
        laporan.jumlah_hadir = jumlah_hadir
        laporan.jumlah_tidak_hadir = jumlah_tidak_hadir
        laporan.bawa_data_kehadiran = bawa_data_kehadiran
        laporan.daftar_hadir = daftar_hadir if bawa_data_kehadiran else None
        laporan.daftar_tidak_hadir = daftar_tidak_hadir if bawa_data_kehadiran else None
        laporan.waktu_input = _now_app_naive()
        message = "Laporan mengajar diperbarui"
        status_code = 200
    else:
        laporan = LaporanMengajar(
            id_monitor=monitor.id_monitor,
            materi=materi,
            catatan=catatan,
            jumlah_hadir=jumlah_hadir,
            jumlah_tidak_hadir=jumlah_tidak_hadir,
            bawa_data_kehadiran=bawa_data_kehadiran,
            daftar_hadir=daftar_hadir if bawa_data_kehadiran else None,
            daftar_tidak_hadir=daftar_tidak_hadir if bawa_data_kehadiran else None,
            waktu_input=_now_app_naive()
        )
        db.session.add(laporan)
        message = "Laporan mengajar tersimpan"
        status_code = 201

    db.session.commit()

    return jsonify({
        "message": message,
        "laporan": _laporan_payload(laporan),
        "sudah_laporan": True
    }), status_code


@monitoring_bp.route("/guru/laporan-mengajar/<int:id_monitor>", methods=["GET"])
@jwt_required()
def get_laporan_mengajar_guru(id_monitor):
    claims = get_jwt()
    if claims.get("role") != "guru":
        return jsonify({"message": "Akses ditolak"}), 403

    id_guru = claims.get("id_guru")
    monitor = LaporanMonitoring.query.get_or_404(id_monitor)

    if not jadwal_milik_guru(monitor.id_jadwal, id_guru):
        return jsonify({"message": "Bukan laporan anda"}), 403

    laporan = LaporanMengajar.query.filter_by(id_monitor=id_monitor).first()

    return jsonify({
        "sudah_laporan": laporan is not None,
        "laporan": _laporan_payload(laporan)
    }), 200


# =====================================================
# ABSEN KELUAR - DITAHAN JIKA BELUM ADA LAPORAN MENGAJAR
# - Update jam keluar di laporan_monitoring
# - Update keterangan rekap harian di kehadiran_guru
# =====================================================

@monitoring_bp.route("/guru/laporan-mengajar/prefill-kehadiran/<int:id_jadwal>", methods=["GET"])
@jwt_required()
def prefill_kehadiran_laporan_mengajar(id_jadwal):
    claims = get_jwt()
    if claims.get("role") != "guru":
        return jsonify({"message": "Akses ditolak"}), 403

    id_guru = claims.get("id_guru")
    if not jadwal_milik_guru(id_jadwal, id_guru):
        return jsonify({"message": "Jadwal tidak valid"}), 403

    jadwal = Jadwal.query.get_or_404(id_jadwal)
    if _jadwal_kelas_selesai(jadwal):
        return jsonify({"message": "Jadwal sudah selesai dan tidak aktif lagi"}), 404

    pertemuan = request.args.get("pertemuan")
    if pertemuan is not None and str(pertemuan).strip() != "":
        try:
            pertemuan = int(pertemuan)
        except Exception:
            return jsonify({"message": "Pertemuan harus berupa angka"}), 400

    return jsonify(_absensi_murid_payload(id_jadwal, pertemuan=pertemuan)), 200


@monitoring_bp.route("/guru/absen-keluar", methods=["POST"])
@jwt_required()
def absen_keluar():
    claims = get_jwt()
    if claims.get("role") != "guru":
        return jsonify({"message": "Akses ditolak"}), 403

    data = request.json or {}
    id_monitor = data.get("id_monitor")
    if not id_monitor:
        return jsonify({"message": "id_monitor wajib"}), 400

    id_guru = claims.get("id_guru")
    monitor = LaporanMonitoring.query.get_or_404(id_monitor)

    if not jadwal_milik_guru(monitor.id_jadwal, id_guru):
        return jsonify({"message": "Bukan data anda"}), 403

    jadwal = Jadwal.query.get(monitor.id_jadwal)
    if _jadwal_kelas_selesai(jadwal):
        return jsonify({"message": "Jadwal sudah selesai dan tidak aktif lagi"}), 400

    if monitor.jam_keluar:
        return jsonify({"message": "Sudah absen keluar"}), 409

    laporan = LaporanMengajar.query.filter_by(id_monitor=monitor.id_monitor).first()
    if not laporan:
        return jsonify({
            "message": "Isi laporan mengajar terlebih dahulu sebelum absen keluar",
            "wajib_laporan": True
        }), 400

    now_time = _now_app().time()
    monitor.jam_keluar = now_time
    monitor.status = "Selesai"

    _upsert_kehadiran_guru(
        id_guru=id_guru,
        tanggal=monitor.tanggal,
        id_jadwal=monitor.id_jadwal,
        status="Hadir",
        keterangan=f"Keluar {now_time.strftime('%H:%M:%S')} - {_jadwal_label(monitor.id_jadwal)}"
    )

    db.session.commit()

    return jsonify({"message": "Absen keluar berhasil"}), 200


# =====================================================
# STATUS ABSEN UNTUK UI GURU
# =====================================================
@monitoring_bp.route("/guru/status-absen/<int:id_jadwal>", methods=["GET"])
@jwt_required()
def status_absen(id_jadwal):
    claims = get_jwt()
    if claims.get("role") != "guru":
        return jsonify({"message": "Akses ditolak"}), 403

    id_guru = claims.get("id_guru")
    if not jadwal_milik_guru(id_jadwal, id_guru):
        return jsonify({"message": "Jadwal tidak valid"}), 403

    today = _today_app()
    jadwal = Jadwal.query.get_or_404(id_jadwal)
    if _jadwal_kelas_selesai(jadwal):
        return jsonify({"message": "Jadwal sudah selesai dan tidak aktif lagi"}), 404

    data = LaporanMonitoring.query.filter_by(
        id_jadwal=id_jadwal,
        tanggal=today
    ).first()

    kehadiran_guru = _get_kehadiran_guru(id_guru, today, id_jadwal)

    if not data:
        status_ui = _status_monitoring(jadwal, None, None, kehadiran_guru)

        # Jika jadwal sudah selesai dan guru tidak absen, simpan Alpa per jadwal
        # agar database sama dengan data yang tampil/download admin.
        if status_ui == "Alpa":
            _upsert_kehadiran_guru(
                id_guru=id_guru,
                tanggal=today,
                id_jadwal=id_jadwal,
                status="Alpa",
                keterangan=f"Auto alpa setelah jadwal selesai - {_jadwal_label(id_jadwal)}"
            )
            db.session.commit()

        kehadiran_guru = _get_kehadiran_guru(id_guru, today, id_jadwal)
        return jsonify({
            "status": status_ui.lower().replace(" ", "_"),
            "status_label": status_ui,
            "id_kehadiran_guru": kehadiran_guru.id_kehadiran if kehadiran_guru else None,
            "alasan": getattr(kehadiran_guru, "alasan", None) if kehadiran_guru else None,
            "instruksi": getattr(kehadiran_guru, "instruksi", None) if kehadiran_guru else None,
            "bukti": getattr(kehadiran_guru, "bukti", None) if kehadiran_guru else None,
            "bukti_url": _bukti_url(getattr(kehadiran_guru, "bukti", None)) if kehadiran_guru else None,
            "status_pengajuan": getattr(kehadiran_guru, "status_pengajuan", None) if kehadiran_guru else None,
            "keterangan": getattr(kehadiran_guru, "keterangan", None) if kehadiran_guru else None,
            "waktu_pengajuan": _waktu_dari_keterangan(
                getattr(kehadiran_guru, "keterangan", None), "pengajuan"
            ) if kehadiran_guru else None,
            "waktu_respon_admin": _waktu_dari_keterangan(
                getattr(kehadiran_guru, "keterangan", None), "respon"
            ) if kehadiran_guru else None,
            "jam_masuk": None,
            "jam_keluar": None,
            "sudah_laporan": False,
            "laporan_mengajar": None
        }), 200

    laporan = LaporanMengajar.query.filter_by(id_monitor=data.id_monitor).first()
    status_ui = _status_monitoring(jadwal, data, laporan, kehadiran_guru)

    if status_ui == "Hadir":
        status_key = "masuk"
    elif status_ui == "Selesai":
        status_key = "keluar"
    else:
        status_key = status_ui.lower().replace(" ", "_")

    return jsonify({
        "status": status_key,
        "status_label": status_ui,
        "id_monitor": data.id_monitor,
        "id_kehadiran_guru": kehadiran_guru.id_kehadiran if kehadiran_guru else None,
        "alasan": getattr(kehadiran_guru, "alasan", None) if kehadiran_guru else None,
        "instruksi": getattr(kehadiran_guru, "instruksi", None) if kehadiran_guru else None,
        "bukti": getattr(kehadiran_guru, "bukti", None) if kehadiran_guru else None,
        "bukti_url": _bukti_url(getattr(kehadiran_guru, "bukti", None)) if kehadiran_guru else None,
        "status_pengajuan": getattr(kehadiran_guru, "status_pengajuan", None) if kehadiran_guru else None,
        "keterangan": getattr(kehadiran_guru, "keterangan", None) if kehadiran_guru else None,
        "waktu_pengajuan": _waktu_dari_keterangan(
            getattr(kehadiran_guru, "keterangan", None), "pengajuan"
        ) if kehadiran_guru else None,
        "waktu_respon_admin": _waktu_dari_keterangan(
            getattr(kehadiran_guru, "keterangan", None), "respon"
        ) if kehadiran_guru else None,
        "jam_masuk": _fmt_time(data.jam_masuk),
        "jam_keluar": _fmt_time(data.jam_keluar),
        "sudah_laporan": laporan is not None,
        "laporan_mengajar": _laporan_payload(laporan)
    }), 200


@monitoring_bp.route("/admin/kehadiran-guru/<int:id_kehadiran>/pengajuan", methods=["POST"])
@jwt_required()
def proses_pengajuan_kehadiran_guru(id_kehadiran):
    claims = get_jwt()
    if claims.get("role") != "admin":
        return jsonify({"message": "Akses ditolak"}), 403

    data = request.json or {}
    aksi = str(data.get("aksi") or data.get("action") or "").strip().lower()
    if aksi not in ["setujui", "approve", "disetujui", "tolak", "reject", "ditolak"]:
        return jsonify({"message": "Aksi harus setujui atau tolak"}), 400

    item = KehadiranGuru.query.get_or_404(id_kehadiran)
    status_manual = _status_kehadiran_manual(item)
    if status_manual not in ["Izin", "Sakit"]:
        return jsonify({"message": "Data ini bukan pengajuan izin/sakit"}), 400

    jadwal = Jadwal.query.get(item.id_jadwal)

    waktu_respon = _now_app().time().strftime("%H:%M:%S")
    if aksi in ["setujui", "approve", "disetujui"]:
        item.status_pengajuan = "Disetujui"
        _add_keterangan(
            item,
            f"Pengajuan {status_manual} disetujui admin {waktu_respon}",
        )
        message = f"Pengajuan {status_manual.lower()} disetujui"
    else:
        item.status_pengajuan = "Ditolak"
        _add_keterangan(
            item,
            f"Pengajuan {status_manual} ditolak admin {waktu_respon}",
        )
        if _jadwal_sudah_selesai(jadwal, item.tanggal):
            item.status = "Alpa"
            _add_keterangan(item, "Menjadi Alpa karena jadwal sudah selesai")
        message = f"Pengajuan {status_manual.lower()} ditolak"

    db.session.commit()

    return jsonify({
        "message": message,
        "id_kehadiran_guru": item.id_kehadiran,
        "status": item.status,
        "status_pengajuan": item.status_pengajuan,
        "alasan": item.alasan,
        "instruksi": item.instruksi,
        "bukti": item.bukti,
        "bukti_url": _bukti_url(item.bukti),
        "waktu_respon_admin": waktu_respon,
        "keterangan": item.keterangan,
    }), 200


# =====================================================
# RIWAYAT MONITORING GURU (MAKSIMAL 1 MINGGU)
# =====================================================
@monitoring_bp.route("/guru/monitoring", methods=["GET"])
@jwt_required()
def monitoring_guru():
    claims = get_jwt()
    if claims.get("role") != "guru":
        return jsonify({"message": "Akses ditolak"}), 403

    if request.args.get("mode", "history").strip().lower() != "history":
        return jsonify({"message": "Mode yang tersedia adalah history"}), 400

    try:
        id_guru = int(claims.get("id_guru"))
    except Exception:
        return jsonify({"message": "id_guru tidak ada di token"}), 400

    today = _today_app()
    cutoff = today - timedelta(days=7)

    # Kehadiran/status tetap memakai tanggal aslinya, sedangkan kelas, mapel,
    # hari, dan jam selalu diambil dari tabel jadwal TERBARU. Karena query
    # melakukan inner join ke Jadwal, jadwal yang dihapus tidak ikut tampil.
    rows = (
        db.session.query(
            Jadwal,
            Kelas,
            MataPelajaran,
            Guru,
            LaporanMonitoring,
            LaporanMengajar,
            KehadiranGuru,
        )
        .join(Jadwal, Jadwal.id_jadwal == KehadiranGuru.id_jadwal)
        .join(Guru, Guru.id_guru == KehadiranGuru.id_guru)
        .join(Kelas, Kelas.id_kelas == Jadwal.id_kelas)
        .join(MataPelajaran, MataPelajaran.id_mapel == Jadwal.id_mapel)
        .outerjoin(
            LaporanMonitoring,
            db.and_(
                LaporanMonitoring.id_jadwal == KehadiranGuru.id_jadwal,
                LaporanMonitoring.tanggal == KehadiranGuru.tanggal,
            ),
        )
        .outerjoin(
            LaporanMengajar,
            LaporanMengajar.id_monitor == LaporanMonitoring.id_monitor,
        )
        .filter(
            KehadiranGuru.id_guru == id_guru,
            KehadiranGuru.tanggal >= cutoff,
            KehadiranGuru.tanggal < today,
            KehadiranGuru.status.in_(["Hadir", "Izin", "Sakit", "Alpa"]),
            _jadwal_kelas_belum_selesai_expr(),
        )
        .order_by(
            KehadiranGuru.tanggal.desc(),
            Jadwal.jam_mulai.desc(),
            Jadwal.id_jadwal.desc(),
        )
        .all()
    )

    return jsonify([
        _monitoring_payload_from_row(row)
        for row in rows
    ]), 200


# =====================================================
# ADMIN MONITORING
# mode=today   -> jadwal hari ini, termasuk yang belum absen
# mode=history -> riwayat monitoring yang sudah tersimpan semua tanggal
# =====================================================
@monitoring_bp.route("/admin/monitoring", methods=["GET"])
@jwt_required()
def monitoring_admin():
    claims = get_jwt()
    if claims.get("role") != "admin":
        return jsonify({"message": "Akses ditolak"}), 403

    mode = request.args.get("mode", "today")
    tanggal_from = request.args.get("from")
    tanggal_to = request.args.get("to")

    _hapus_monitoring_lebih_14_hari()
    data = _query_monitoring_rows(mode, tanggal_from, tanggal_to)
    _sinkron_kehadiran_guru_terjadwal(data)

    return jsonify([
        _monitoring_payload_from_row(row)
        for row in data
    ]), 200


# =====================================================
# HELPER EXPORT MONITORING
# =====================================================
def _kehadiran_excel(payload):
    """
    Kolom Kehadiran berisi status hadir/tidak hadir guru per jadwal/mapel.
    Bukan lagi rekap harian guru, supaya 1 guru yang mengampu 2 mapel
    bisa tampil: mapel A Hadir, mapel B Alpa.
    """
    status = str(payload.get("status") or "").strip().lower()
    if status in ["selesai", "hadir"]:
        return "Hadir"
    if status == "izin":
        return "Izin"
    if status == "sakit":
        return "Sakit"
    if status == "alpa":
        return "Alpa"

    raw_kehadiran = payload.get("kehadiran_guru")
    if raw_kehadiran is not None:
        text = str(raw_kehadiran).strip().lower()
        if text in ["hadir", "masuk", "selesai"]:
            return "Hadir"
        if text in ["izin", "ijin"]:
            return "Izin"
        if text == "sakit":
            return "Sakit"
        if text in ["alpa", "alpha", "tidak hadir", "tidak_hadir"]:
            return "Alpa"

    if payload.get("masuk"):
        return "Hadir"

    return "Alpa"


def _alasan_instruksi_excel(payload):
    alasan = str(payload.get("alasan") or "").strip()
    instruksi = str(payload.get("instruksi") or "").strip()
    bagian = []
    if alasan:
        bagian.append(alasan.rstrip(" .") + ".")
    if instruksi:
        bagian.append(
            "Instruksi yang telah diberikan kepada murid/guru pengganti "
            + instruksi.lstrip()
        )
    return " ".join(bagian).strip() or "-"


def _status_excel(payload):
    """
    Kolom Status tetap mengambil status proses monitoring.
    Contoh: Belum Absen, Hadir, Selesai.
    """
    status = str(payload.get("status") or "").strip()
    return status if status else "Belum Absen"


# =====================================================
# ADMIN DOWNLOAD XLSX MONITORING GURU
# Hasil: NIP | Nama Guru | Mapel | Kehadiran | Tanggal | Status
# =====================================================
@monitoring_bp.route("/admin/monitoring/export", methods=["GET"])
@jwt_required()
def export_monitoring_admin():
    claims = get_jwt()
    if claims.get("role") != "admin":
        return jsonify({"message": "Akses ditolak"}), 403

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        return jsonify({
            "message": "Library openpyxl belum terpasang. Jalankan: pip install openpyxl"
        }), 500

    mode = request.args.get("mode", "today")
    tanggal_from = request.args.get("from")
    tanggal_to = request.args.get("to")

    _hapus_monitoring_lebih_14_hari()
    data = _query_monitoring_rows(mode, tanggal_from, tanggal_to)
    _sinkron_kehadiran_guru_terjadwal(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoring Guru"

    header = [
        "NIP",
        "Nama Guru",
        "Mapel",
        "Kehadiran",
        "Tanggal",
        "Status",
        "Alasan",
        "Murid Hadir",
        "Materi",
        "Murid Tidak Hadir",
        "NIS/Nama Hadir",
        "NIS/Nama Tidak Hadir",
    ]
    ws.append(header)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A6AA1")
        cell.alignment = Alignment(horizontal="center")

    for row in data:
        payload = _monitoring_payload_from_row(row)
        laporan_payload = payload.get("laporan_mengajar") or {}
        ws.append([
            _safe_sheet_value(payload.get("nip")),
            _safe_sheet_value(payload.get("nama_guru") or payload.get("guru")),
            _safe_sheet_value(payload.get("mapel")),
            _safe_sheet_value(_kehadiran_excel(payload)),
            _safe_sheet_value(payload.get("tanggal")),
            _safe_sheet_value(_status_excel(payload)),
            _safe_sheet_value(_alasan_instruksi_excel(payload)),
            _safe_sheet_value(
                payload.get("jumlah_hadir")
                if payload.get("jumlah_hadir") is not None
                else laporan_payload.get("jumlah_hadir")
            ),
            _safe_sheet_value(payload.get("materi") or laporan_payload.get("materi")),
            _safe_sheet_value(
                payload.get("jumlah_tidak_hadir")
                if payload.get("jumlah_tidak_hadir") is not None
                else laporan_payload.get("jumlah_tidak_hadir")
            ),
            _safe_sheet_value(payload.get("daftar_hadir") or laporan_payload.get("daftar_hadir")),
            _safe_sheet_value(payload.get("daftar_tidak_hadir") or laporan_payload.get("daftar_tidak_hadir")),
        ])

    widths = {
        "A": 18,
        "B": 28,
        "C": 28,
        "D": 18,
        "E": 16,
        "F": 18,
        "G": 54,
        "H": 16,
        "I": 34,
        "J": 20,
        "K": 42,
        "L": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    suffix = mode
    if tanggal_from or tanggal_to:
        suffix += f"_{tanggal_from or 'awal'}_{tanggal_to or 'akhir'}"

    filename = f"monitoring_guru_{suffix}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
